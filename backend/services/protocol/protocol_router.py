from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from database import get_db
from schemas.protocol import ProtocolCreate, ProtocolUpdate, ProtocolResponse
from services.protocol.protocol_service import ProtocolService
from config.logging import logger
import openpyxl
from io import BytesIO
import json
from datetime import datetime

router = APIRouter(
    prefix="/api/protocols",
    tags=["protocol"],
    responses={404: {"description": "Not found"}},
)


@router.get("", response_model=List[ProtocolResponse])
def get_protocols(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    """获取协议列表"""
    logger.info(f"获取协议列表，skip: {skip}, limit: {limit}")
    protocols = ProtocolService.get_protocols(db, skip=skip, limit=limit)
    logger.info(f"获取到 {len(protocols)} 个协议")
    return protocols


@router.get("/export")
def export_protocols(db: Session = Depends(get_db)):
    """导出协议到 Excel"""
    logger.info("导出协议到 Excel")

    # 获取所有协议
    protocols = ProtocolService.get_protocols(db, skip=0, limit=1000)

    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "协议列表"

    # 写入表头
    headers = ['协议 ID', '协议名称', '协议类型', '设备类型', '交互对象', '协议版本', '协议描述', '字段说明']
    ws.append(headers)

    # 写入数据
    for protocol in protocols:
        # 解析 fields 字段
        fields = ''
        if protocol.fields and isinstance(protocol.fields, list):
            fields = '\n'.join([f"{field['name']}: {field['type']} ({field['required']})" for field in protocol.fields])

        row = [
            protocol.id,
            protocol.name,
            protocol.type,
            protocol.device_type,
            protocol.interaction_object,
            protocol.version,
            protocol.description or '',
            fields
        ]
        ws.append(row)

    # 调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"导出完成，共 {len(protocols)} 个协议")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=protocols_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )


@router.post("/import")
def import_protocols(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入协议从 Excel"""
    logger.info("导入协议从 Excel")

    # 检查文件类型
    if not file.filename.endswith('.xlsx') and not file.filename.endswith('.xls'):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件 (.xlsx, .xls)")

    # 读取文件
    try:
        wb = openpyxl.load_workbook(file.file)
        ws = wb.active
    except Exception as e:
        logger.error(f"解析 Excel 文件失败：{e}")
        raise HTTPException(status_code=400, detail="解析 Excel 文件失败")

    # 读取数据
    success_count = 0
    failed_count = 0

    # 跳过表头
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            # 解析字段
            protocol_id = row[0] or f"IMPORT-{datetime.now().strftime('%Y%m%d%H%M%S')}-{success_count + failed_count}"
            name = row[1] or ""
            type_ = row[2] or "system_message"
            device_type = row[3] or "Host"
            interaction_object = row[4] or "Superior"
            version = row[5] or "1.0"
            description = row[6] or ""
            fields_str = row[7] or ""

            # 解析字段说明
            fields = []
            if fields_str:
                field_lines = fields_str.split('\n')
                for line in field_lines:
                    if line.strip():
                        parts = line.split(': ')
                        if len(parts) == 2:
                            field_name = parts[0].strip()
                            field_details = parts[1].strip()
                            field_type = field_details.split(' (')[0]
                            required = field_details.split(' (')[1].rstrip(')')
                            fields.append({
                                'name': field_name,
                                'type': field_type,
                                'length': '',
                                'required': required,
                                'description': ''
                            })

            # 创建协议
            protocol = ProtocolCreate(
                id=protocol_id,
                name=name,
                type=type_,
                device_type=device_type,
                interaction_object=interaction_object,
                version=version,
                structure=f"<?xml version=\"1.0\" encoding=\"UTF-8\"?><Protocol><SendCode>sender</SendCode><ReceiveCode>receiver</ReceiveCode><Type>message</Type><Code>0000</Code><Command>command</Command><Time>timestamp</Time></Protocol>",
                fields=fields,
                description=description
            )

            # 保存到数据库
            ProtocolService.create_protocol(db, protocol)
            success_count += 1
        except Exception as e:
            logger.error(f"导入协议失败：{e}")
            failed_count += 1

    logger.info(f"导入完成，成功 {success_count} 个，失败 {failed_count} 个")

    return {
        "success": success_count,
        "failed": failed_count
    }


# ====== 注意：具体路由必须在参数路由之前注册 ======

@router.get("/grouped")
def get_protocols_grouped(db: Session = Depends(get_db)):
    """获取按设备类型和消息类型分组的协议列表"""
    logger.info("获取分组协议列表")
    grouped = ProtocolService.get_protocols_grouped(db)
    return grouped


@router.get("/dictionary/device-types")
def get_device_types(db: Session = Depends(get_db)):
    """获取设备类型字典（带中文映射）"""
    from services.dictionary.dictionary_service import DictionaryService
    
    logger.info("获取设备类型字典")
    # 获取设备类型字典
    device_type_dict = DictionaryService.get_dictionaries_by_type(db, dictionary_type="device_type")
    
    # 转换为前端需要的格式
    result = {}
    for item in device_type_dict:
        result[item.code] = {
            "code": item.code,
            "name": item.name,
            "description": item.description
        }
    
    logger.info(f"获取到 {len(result)} 个设备类型")
    return result


@router.get("/dictionary/protocol-types")
def get_protocol_types(db: Session = Depends(get_db)):
    """获取协议类型字典（带中文映射）"""
    from services.dictionary.dictionary_service import DictionaryService
    
    logger.info("获取协议类型字典")
    # 获取协议类型字典
    protocol_type_dict = DictionaryService.get_dictionaries_by_type(db, dictionary_type="protocol_type")
    
    # 转换为前端需要的格式
    result = {}
    for item in protocol_type_dict:
        result[item.code] = {
            "code": item.code,
            "name": item.name,
            "description": item.description
        }
    
    logger.info(f"获取到 {len(result)} 个协议类型")
    return result


@router.get("/{protocol_id}", response_model=ProtocolResponse)
def get_protocol(protocol_id: str, db: Session = Depends(get_db)):
    """根据 ID 获取协议"""
    logger.info(f"根据 ID 获取协议：{protocol_id}")
    protocol = ProtocolService.get_protocol(db, protocol_id=protocol_id)
    if protocol is None:
        logger.warning(f"协议未找到：{protocol_id}")
        raise HTTPException(status_code=404, detail="Protocol not found")
    logger.info(f"获取协议成功：{protocol_id}")
    return protocol


@router.post("", response_model=ProtocolResponse)
def create_protocol(protocol: ProtocolCreate, db: Session = Depends(get_db)):
    """创建协议"""
    logger.info(f"创建协议请求：{protocol.model_dump()}")
    created_protocol = ProtocolService.create_protocol(db=db, protocol=protocol)
    logger.info(f"协议创建成功：{created_protocol.id}, 响应：{created_protocol.model_dump()}")
    return created_protocol


@router.put("/{protocol_id}", response_model=ProtocolResponse)
def update_protocol(protocol_id: str, protocol: ProtocolUpdate, db: Session = Depends(get_db)):
    """更新协议"""
    logger.info(f"更新协议请求：protocol_id={protocol_id}, data={protocol.model_dump(exclude_unset=True)}")
    db_protocol = ProtocolService.update_protocol(db=db, protocol_id=protocol_id, protocol_update=protocol)
    if db_protocol is None:
        logger.warning(f"协议未找到：{protocol_id}")
        raise HTTPException(status_code=404, detail="Protocol not found")
    logger.info(f"协议更新成功：{protocol_id}, 响应：{db_protocol.model_dump()}")
    return db_protocol


@router.delete("/{protocol_id}")
def delete_protocol(protocol_id: str, db: Session = Depends(get_db)):
    """删除协议"""
    logger.info(f"删除协议：{protocol_id}")
    success = ProtocolService.delete_protocol(db=db, protocol_id=protocol_id)
    if not success:
        logger.warning(f"协议未找到：{protocol_id}")
        raise HTTPException(status_code=404, detail="Protocol not found")
    logger.info(f"协议删除成功：{protocol_id}")
    return {"message": "Protocol deleted successfully"}


@router.get("/device/{device_type}", response_model=List[ProtocolResponse])
def get_protocols_by_device_type(device_type: str, db: Session = Depends(get_db)):
    """根据设备类型获取协议列表"""
    logger.info(f"根据设备类型获取协议列表：{device_type}")
    protocols = ProtocolService.get_protocols_by_device_type(db, device_type=device_type)
    logger.info(f"获取到 {len(protocols)} 个协议")
    return protocols


@router.get("/type/{protocol_type}", response_model=List[ProtocolResponse])
def get_protocols_by_type(protocol_type: str, db: Session = Depends(get_db)):
    """根据协议类型获取协议列表"""
    logger.info(f"根据协议类型获取协议列表：{protocol_type}")
    protocols = ProtocolService.get_protocols_by_type(db, protocol_type=protocol_type)
    logger.info(f"获取到 {len(protocols)} 个协议")
    return protocols
